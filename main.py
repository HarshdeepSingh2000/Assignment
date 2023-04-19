import openpyxl

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set the headers for the columns
ws['A1'] = "Rank"
ws['B1'] = "Name"
ws['C1'] = "UID"
ws['D1'] = "No. of Statements"
ws['E1'] = "No. of Reasons"

# Insert the data into the worksheet
data = [
    [1, "Ronak 0", 302, 16, 19],
    [2, "Soumita M", 280, 13, 21],
    [3, "Subhangi 0", 75, 13, 16],
    [4, "Merwin", 295, 13, 12],
    [5, "Nitin Shane", 1124, 12, 11],
    [6, "Ayisha", 3406, 11, 8],
    [7, "imshawan", 71, 9, 9],
    [8, "sharath", 3367, 8, 8],
    [9, "Anuraj_Saini", 342, 7, 7],
    [10, "devmenkr", 360, 7, 7],
    [11, "Saurabh", 271, 6, 8],
    [12, "Palash", 69, 6, 7],
    [13, "raman", 539, 5, 5],
    [14, "Nishant", 299, 4, 5],
    [15, "_riddhi_213", 504, 2, 7],
    [16, "Amrit Malviya", 336, 2, 4],
    [17, "darshimalde", 3169, 3, 3],
    [18, "Rohit Dutta", 266, 3, 3],
    [19, "fardinkamal62", 263, 2, 2],
    [20, "Shagun", 100, 2, 2],
    [21, "Vatsal", 3408, 2, 2]
]

for row in data:
    ws.append(row)

# Save the workbook to a file
wb.save("rankings.xlsx")
